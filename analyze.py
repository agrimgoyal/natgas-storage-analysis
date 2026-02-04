"""
Natural Gas Storage Analysis - Enhanced Edition (v4.2)
Updates:
- Added plots for 5-year moving averages
- Added monthly/weekly average storage and withdrawal/injection charts
- Enhanced color coding and formatting
- Improved visual hierarchy
- Using matplotlib for better charts with axis labels
- Improved chart UI and readability
"""

import pandas as pd  # type: ignore
import numpy as np  # type: ignore
import calendar  # type: ignore
from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # type: ignore
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule  # type: ignore
from openpyxl.utils import get_column_letter  # type: ignore
from openpyxl.drawing.image import Image as XLImage  # type: ignore
import matplotlib  # type: ignore

matplotlib.use("Agg")  # Use non-interactive backend
import matplotlib.pyplot as plt  # type: ignore
import matplotlib.dates as mdates  # type: ignore
from datetime import datetime  # type: ignore
import io  # type: ignore
import warnings  # type: ignore
import win32com.client  # type: ignore
import os  # type: ignore

warnings.filterwarnings("ignore")

# Set matplotlib style with improved readability
plt.rcParams["font.family"] = "sans-serif"
plt.rcParams["font.sans-serif"] = ["Arial", "DejaVu Sans", "Helvetica"]
plt.rcParams["font.size"] = 10
plt.rcParams["axes.labelsize"] = 11
plt.rcParams["axes.titlesize"] = 13
plt.rcParams["xtick.labelsize"] = 9
plt.rcParams["ytick.labelsize"] = 9
plt.rcParams["legend.fontsize"] = 10
plt.rcParams["figure.titlesize"] = 14
plt.rcParams["axes.grid"] = True
plt.rcParams["grid.alpha"] = 0.3
plt.rcParams["grid.linestyle"] = "--"
plt.rcParams["axes.facecolor"] = "#F8F9FA"
plt.rcParams["figure.facecolor"] = "white"


class NaturalGasStorageAnalyzer:
    def __init__(self, file_path):
        self.file_path = file_path
        self.df_master = None
        self.regions = {}

        # Enhanced color scheme
        self.colors = {
            "header": "1F4E78",  # Deep blue
            "subheader": "2E75B6",  # Medium blue
            "positive": "70AD47",  # Green
            "negative": "C00000",  # Red
            "neutral": "FFC000",  # Orange
            "background": "F2F2F2",  # Light gray
            "current": "4472C4",  # Blue
            "historical": "7F7F7F",  # Gray
            "avg_line": "FF6600",  # Orange
        }

        # Regional storage capacity in BCF (based on EIA data)
        # Source: EIA Natural Gas Storage Capacity
        # https://www.eia.gov/naturalgas/storagecapacity/
        self.regional_capacity = {
            "Total_Lower_48": 4671,
            "East": 1043,
            "Midwest": 1221,
            "Mountain": 479,
            "Pacific": 369,
            "South_Central": 1558,
            "Salt": 471,
            "Non_Salt": 1088,
        }

    def load_and_clean_data(self):
        print(f"Loading data from: {self.file_path}")
        # Load the raw file without a header first to locate the real header
        df_raw = pd.read_excel(self.file_path, header=None)

        header_row_idx = None
        # Scan the first 20 rows to find the header
        for idx in range(min(20, len(df_raw))):
            # Robust conversion: Force every value to string individually using list comprehension
            # This prevents the "float is not iterable" error even if pandas inference fails
            row_vals = [str(v).lower() for v in df_raw.iloc[idx]]

            if any("week ending" in v for v in row_vals) and any(
                "total lower 48" in v for v in row_vals
            ):
                header_row_idx = idx
                break

        if header_row_idx is None:
            raise ValueError("Could not find header row.")

        # Reload with the correct header
        df = pd.read_excel(self.file_path, header=header_row_idx)

        # Identify the date column dynamically
        date_col = next(
            (c for c in df.columns if "week ending" in str(c).lower()), None
        )
        if date_col is None:
            raise ValueError("Could not find 'Week Ending' column.")

        df = df.rename(columns={date_col: "Week_Ending"})
        df = df.dropna(subset=["Week_Ending"])
        df["Week_Ending"] = pd.to_datetime(df["Week_Ending"], errors="coerce")
        df = (
            df.dropna(subset=["Week_Ending"])
            .sort_values("Week_Ending")
            .reset_index(drop=True)
        )

        target_keywords = {
            "Total_Lower_48": ["total lower 48"],
            "Salt": ["salt", "south central salt"],
            "Non_Salt": ["nonsalt", "non-salt", "south central nonsalt"],
            "East": ["east"],
            "Midwest": ["midwest"],
            "Mountain": ["mountain"],
            "Pacific": ["pacific"],
            "South_Central": ["south central"],
        }

        # Dynamic column mapping
        for col in df.columns:
            col_lower = str(col).lower().strip()
            if col == "Week_Ending":
                continue
            for region_key, keywords in target_keywords.items():
                if region_key == "South_Central" and ("salt" in col_lower):
                    continue
                if any(k == col_lower or k in col_lower for k in keywords):
                    if region_key not in self.regions:
                        self.regions[region_key] = col
                        df[col] = pd.to_numeric(df[col], errors="coerce")
                    break

        self.df_master = df
        return df

    def analyze_dataset(self, df_input, value_column, freq="Weekly"):
        # Prepare Data
        df = df_input[["Week_Ending", value_column]].copy()
        df = df.rename(columns={value_column: "Storage_Level"})
        df = df.dropna().sort_values("Week_Ending").reset_index(drop=True)

        df["Year"] = df["Week_Ending"].dt.year
        df["Week_Num"] = df["Week_Ending"].dt.isocalendar().week
        df["Month"] = df["Week_Ending"].dt.month
        df["Net_Flow"] = df["Storage_Level"].diff()

        # Categorize flow type
        df["Flow_Type"] = df["Net_Flow"].apply(
            lambda x: "Injection" if x > 0 else ("Withdrawal" if x < 0 else "Flat")
        )

        # Fast Lookup Index
        df_lookup = df.set_index("Week_Ending").sort_index()

        avg_flows = [np.nan] * len(df)
        avg_levels = [np.nan] * len(df)
        diff_flows = [np.nan] * len(df)
        diff_levels = [np.nan] * len(df)
        yoy_flows = [np.nan] * len(df)

        tol = pd.Timedelta(days=4) if freq == "Weekly" else pd.Timedelta(days=15)

        for i, current_date in enumerate(df["Week_Ending"]):
            hist_vals_flow = []
            hist_vals_level = []

            for offset in range(1, 6):  # 1 to 5 years back
                target_year = current_date.year - offset
                try:
                    target_date = current_date.replace(year=target_year)
                except ValueError:
                    target_date = current_date.replace(year=target_year, day=28)

                start_search = target_date - tol
                end_search = target_date + tol

                try:
                    matches = df_lookup[start_search:end_search]
                    if not matches.empty:
                        if len(matches) > 1:
                            dates = matches.index.to_series()
                            best_idx = (dates - target_date).abs().idxmin()
                            best_match = matches.loc[best_idx]
                        else:
                            best_match = matches.iloc[0]

                        if pd.notna(best_match["Net_Flow"]):
                            hist_vals_flow.append(best_match["Net_Flow"])
                        if pd.notna(best_match["Storage_Level"]):
                            hist_vals_level.append(best_match["Storage_Level"])
                except KeyError:
                    pass

            if len(hist_vals_flow) >= 3:
                avg = sum(hist_vals_flow) / len(hist_vals_flow)
                avg_flows[i] = avg
                diff_flows[i] = df.at[i, "Net_Flow"] - avg

            if len(hist_vals_level) >= 3:
                avg = sum(hist_vals_level) / len(hist_vals_level)
                avg_levels[i] = avg
                diff_levels[i] = df.at[i, "Storage_Level"] - avg

            # YoY Calculation
            target_year = current_date.year - 1
            try:
                target_date = current_date.replace(year=target_year)
            except:
                target_date = current_date.replace(year=target_year, day=28)
            start_search = target_date - tol
            end_search = target_date + tol

            try:
                matches = df_lookup[start_search:end_search]
                if not matches.empty:
                    if len(matches) > 1:
                        dates = matches.index.to_series()
                        best_idx = (dates - target_date).abs().idxmin()
                        best_match = matches.loc[best_idx]
                    else:
                        best_match = matches.iloc[0]
                    if pd.notna(best_match["Net_Flow"]):
                        yoy_flows[i] = df.at[i, "Net_Flow"] - best_match["Net_Flow"]
            except KeyError:
                pass

        df["5Yr_Avg_Flow"] = avg_flows
        df["Diff_vs_5Yr_Flow"] = diff_flows
        df["5Yr_Avg_Storage"] = avg_levels
        df["Diff_vs_5Yr_Storage"] = diff_levels
        df["YoY_Change_Flow"] = yoy_flows

        return df

    def create_monthly_data(self, df_weekly, value_col):
        df = df_weekly[["Week_Ending", value_col]].copy()
        df["Year"] = df["Week_Ending"].dt.year
        df["Month"] = df["Week_Ending"].dt.month
        return (
            df.sort_values("Week_Ending")
            .groupby(["Year", "Month"])
            .last()
            .reset_index()
            .rename(columns={value_col: "Storage_Level"})
        )

    def auto_resize_columns(self, ws):
        """Robust resizing scanning ALL rows."""
        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)

            if col[0].value:
                max_length = len(str(col[0].value)) * 1.25

            for cell in col:
                try:
                    if cell.value:
                        val = cell.value
                        if isinstance(val, (int, float)):
                            val_len = len(f"{val:,.1f}")
                        elif isinstance(val, pd.Timestamp) or "datetime" in str(
                            type(val)
                        ):
                            val_len = 12
                        else:
                            val_len = len(str(val))

                        if val_len > max_length:
                            max_length = val_len
                except:
                    pass

            adj_width = (max_length + 2) * 1.1
            ws.column_dimensions[column_letter].width = min(max(adj_width, 10), 50)

    def add_moving_average_chart(self, ws, df, start_row, start_col):
        """Add chart showing current storage vs 5-year moving average using matplotlib."""

        # Get recent data (last 2 years)
        recent_df = df.tail(104) if len(df) > 104 else df

        # Create figure with improved styling
        fig, ax = plt.subplots(figsize=(14, 7), facecolor="white")
        ax.set_facecolor("#F8F9FA")

        # Plot data with enhanced styling
        ax.plot(
            recent_df["Week_Ending"],
            recent_df["Storage_Level"],
            color="#2E75B6",
            linewidth=3,
            label="Current Storage",
            marker="o",
            markersize=4,
            markevery=8,
            markeredgewidth=1.5,
            markeredgecolor="white",
            zorder=3,
        )

        ax.plot(
            recent_df["Week_Ending"],
            recent_df["5Yr_Avg_Storage"],
            color="#FF6600",
            linewidth=2.5,
            linestyle="--",
            label="5-Year Average",
            marker="s",
            markersize=4,
            markevery=8,
            markeredgewidth=1.5,
            markeredgecolor="white",
            zorder=3,
        )

        # Formatting
        ax.set_title(
            "Storage Level vs 5-Year Average",
            fontsize=16,
            fontweight="bold",
            pad=20,
            color="#1F4E78",
        )
        ax.set_xlabel("Date", fontsize=13, fontweight="bold", color="#333333")
        ax.set_ylabel("Storage (BCF)", fontsize=13, fontweight="bold", color="#333333")

        # Format x-axis with better date formatting
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %Y"))
        ax.xaxis.set_major_locator(mdates.MonthLocator(interval=3))
        plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha="right", fontsize=10)

        # Format y-axis with commas and better spacing
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f"{int(x):,}"))
        ax.tick_params(axis="both", which="major", labelsize=10, colors="#333333")

        # Enhanced grid
        ax.grid(
            True, alpha=0.4, linestyle="--", linewidth=0.8, color="#CCCCCC", zorder=1
        )
        ax.set_axisbelow(True)

        # Add border around plot area
        for spine in ax.spines.values():
            spine.set_edgecolor("#CCCCCC")
            spine.set_linewidth(1.5)

        # Enhanced legend with better styling
        legend = ax.legend(
            loc="upper left",
            frameon=True,
            shadow=True,
            fontsize=11,
            fancybox=True,
            framealpha=0.95,
            edgecolor="#CCCCCC",
        )
        legend.get_frame().set_facecolor("white")

        # Tight layout
        plt.tight_layout()

        # Save to bytes with high quality
        img_buffer = io.BytesIO()
        plt.savefig(
            img_buffer,
            format="png",
            dpi=150,
            bbox_inches="tight",
            facecolor="white",
            edgecolor="none",
        )
        img_buffer.seek(0)
        plt.close()

        # Insert into Excel
        img = XLImage(img_buffer)
        img.width = 840
        img.height = 420
        ws.add_image(img, f"{get_column_letter(start_col)}{start_row}")

        return start_row + 24  # Return next available row

    def add_injection_withdrawal_chart(self, ws, df, freq, start_row, start_col):
        """Add chart showing average injection/withdrawal patterns using matplotlib."""
        index_col = "Week_Num" if freq == "Weekly" else "Month"

        # Calculate average flows by period
        avg_flows = df.groupby(index_col)["Net_Flow"].mean().reset_index()
        avg_flows.columns = ["Period", "Avg_Flow"]

        # Create period labels
        if freq == "Monthly":
            period_labels = [calendar.month_abbr[int(p)] for p in avg_flows["Period"]]
        else:
            period_labels = avg_flows["Period"].astype(str).tolist()

        # Create figure with improved styling
        fig, ax = plt.subplots(figsize=(14, 7), facecolor="white")
        ax.set_facecolor("#F8F9FA")

        # Color bars based on positive/negative with better colors
        colors = ["#5B9BD5" if x > 0 else "#ED7D31" for x in avg_flows["Avg_Flow"]]

        # Create bar chart
        bars = ax.bar(
            period_labels,
            avg_flows["Avg_Flow"],
            color=colors,
            edgecolor="#333333",
            linewidth=1.2,
            alpha=0.85,
            zorder=3,
        )

        # Add value labels on bars with better positioning
        for i, (bar, val) in enumerate(zip(bars, avg_flows["Avg_Flow"])):
            height = bar.get_height()
            label_y = height + (5 if height > 0 else -5)
            ax.text(
                bar.get_x() + bar.get_width() / 2.0,
                label_y,
                f"{val:.1f}",
                ha="center",
                va="bottom" if height > 0 else "top",
                fontsize=9,
                fontweight="bold",
                color="#333333",
            )

        # Formatting
        ax.set_title(
            f"Average {freq} Injection/Withdrawal Pattern",
            fontsize=16,
            fontweight="bold",
            pad=20,
            color="#1F4E78",
        )
        ax.set_xlabel(
            "Month" if freq == "Monthly" else "Week Number",
            fontsize=13,
            fontweight="bold",
            color="#333333",
        )
        ax.set_ylabel("Net Flow (BCF)", fontsize=13, fontweight="bold", color="#333333")

        # Add zero line with better styling
        ax.axhline(y=0, color="#333333", linestyle="-", linewidth=2, zorder=2)

        # Format y-axis
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f"{int(x):,}"))
        ax.tick_params(axis="both", which="major", labelsize=10, colors="#333333")

        # Enhanced grid
        ax.grid(
            True,
            alpha=0.4,
            linestyle="--",
            linewidth=0.8,
            color="#CCCCCC",
            axis="y",
            zorder=1,
        )
        ax.set_axisbelow(True)

        # Add border around plot area
        for spine in ax.spines.values():
            spine.set_edgecolor("#CCCCCC")
            spine.set_linewidth(1.5)

        # Rotate x labels if weekly
        if freq == "Weekly":
            plt.setp(
                ax.xaxis.get_majorticklabels(), rotation=45, ha="right", fontsize=9
            )
        else:
            plt.setp(ax.xaxis.get_majorticklabels(), fontsize=10)

        # Enhanced legend
        from matplotlib.patches import Patch  # type: ignore

        legend_elements = [
            Patch(
                facecolor="#5B9BD5",
                edgecolor="#333333",
                label="Injection",
                linewidth=1.2,
            ),
            Patch(
                facecolor="#ED7D31",
                edgecolor="#333333",
                label="Withdrawal",
                linewidth=1.2,
            ),
        ]
        legend = ax.legend(
            handles=legend_elements,
            loc="upper right",
            frameon=True,
            shadow=True,
            fontsize=11,
            fancybox=True,
            framealpha=0.95,
            edgecolor="#CCCCCC",
        )
        legend.get_frame().set_facecolor("white")

        # Tight layout
        plt.tight_layout()

        # Save to bytes
        img_buffer = io.BytesIO()
        plt.savefig(
            img_buffer,
            format="png",
            dpi=150,
            bbox_inches="tight",
            facecolor="white",
            edgecolor="none",
        )
        img_buffer.seek(0)
        plt.close()

        # Insert into Excel
        img = XLImage(img_buffer)
        img.width = 840
        img.height = 420
        ws.add_image(img, f"{get_column_letter(start_col)}{start_row}")

        return start_row + 24  # Return next available row

    def add_average_storage_chart(self, ws, df, freq, start_row, start_col):
        """Add chart showing average storage levels by period using matplotlib."""
        index_col = "Week_Num" if freq == "Weekly" else "Month"

        # Calculate average storage by period
        avg_storage = df.groupby(index_col)["Storage_Level"].mean().reset_index()
        avg_storage.columns = ["Period", "Avg_Storage"]

        # Create period labels
        if freq == "Monthly":
            period_labels = [calendar.month_abbr[int(p)] for p in avg_storage["Period"]]
        else:
            period_labels = avg_storage["Period"].astype(str).tolist()

        # Create figure with improved styling
        fig, ax = plt.subplots(figsize=(14, 7), facecolor="white")
        ax.set_facecolor("#F8F9FA")

        # Create line chart with enhanced styling
        ax.plot(
            period_labels,
            avg_storage["Avg_Storage"],
            color="#2E75B6",
            linewidth=3.5,
            marker="o",
            markersize=7,
            markerfacecolor="#2E75B6",
            markeredgecolor="white",
            markeredgewidth=2,
            zorder=3,
        )

        # Fill area under curve with gradient effect
        ax.fill_between(
            range(len(period_labels)),
            avg_storage["Avg_Storage"],
            alpha=0.25,
            color="#2E75B6",
            zorder=2,
        )

        # Add value labels on points with better styling
        skip = 1 if freq == "Monthly" else 4
        for i, (label, val) in enumerate(
            zip(period_labels, avg_storage["Avg_Storage"])
        ):
            if i % skip == 0:
                ax.text(
                    i,
                    val + (max(avg_storage["Avg_Storage"]) * 0.02),
                    f"{val:.0f}",
                    ha="center",
                    va="bottom",
                    fontsize=9,
                    fontweight="bold",
                    color="#1F4E78",
                    bbox=dict(
                        boxstyle="round,pad=0.3",
                        facecolor="white",
                        edgecolor="#CCCCCC",
                        alpha=0.8,
                    ),
                )

        # Formatting
        ax.set_title(
            f"Average {freq} Storage Levels",
            fontsize=16,
            fontweight="bold",
            pad=20,
            color="#1F4E78",
        )
        ax.set_xlabel(
            "Month" if freq == "Monthly" else "Week Number",
            fontsize=13,
            fontweight="bold",
            color="#333333",
        )
        ax.set_ylabel("Storage (BCF)", fontsize=13, fontweight="bold", color="#333333")

        # Format y-axis
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f"{int(x):,}"))
        ax.tick_params(axis="both", which="major", labelsize=10, colors="#333333")

        # Enhanced grid
        ax.grid(
            True, alpha=0.4, linestyle="--", linewidth=0.8, color="#CCCCCC", zorder=1
        )
        ax.set_axisbelow(True)

        # Add border around plot area
        for spine in ax.spines.values():
            spine.set_edgecolor("#CCCCCC")
            spine.set_linewidth(1.5)

        # Rotate x labels if weekly
        if freq == "Weekly":
            plt.setp(
                ax.xaxis.get_majorticklabels(), rotation=45, ha="right", fontsize=9
            )
        else:
            plt.setp(ax.xaxis.get_majorticklabels(), fontsize=10)

        # Tight layout
        plt.tight_layout()

        # Save to bytes
        img_buffer = io.BytesIO()
        plt.savefig(
            img_buffer,
            format="png",
            dpi=150,
            bbox_inches="tight",
            facecolor="white",
            edgecolor="none",
        )
        img_buffer.seek(0)
        plt.close()

        # Insert into Excel
        img = XLImage(img_buffer)
        img.width = 840
        img.height = 420
        ws.add_image(img, f"{get_column_letter(start_col)}{start_row}")

        return start_row + 24  # Return next available row

    def add_seasonality_chart(
        self, ws, df, freq="Weekly", region_name="Total_Lower_48"
    ):
        """Creates Key Stats (Left) -> Pivot Table -> Chart (Right) using matplotlib."""
        index_col = "Week_Num" if freq == "Weekly" else "Month"
        current_year = df["Year"].max()
        df_recent = df[df["Year"] >= current_year - 10]

        pivot = df_recent.pivot_table(
            index=index_col, columns="Year", values="Storage_Level", aggfunc="mean"
        )

        if freq == "Monthly":
            month_map = {i: calendar.month_name[i] for i in range(1, 13)}
            pivot.index = pivot.index.map(month_map)

        # --- LAYOUT DEFINITIONS ---
        stats_start_col = 14  # Column N
        pivot_start_col = 17  # Column Q
        start_row = 3

        # --- 1. WRITE KEY STATS (LEFT) ---
        min_val = pivot.min().min()
        max_val = pivot.max().max()
        avg_val = pivot.mean().mean()

        # Get regional capacity
        capacity_val = self.regional_capacity.get(region_name, None)

        stats_col_let = get_column_letter(stats_start_col)

        # Enhanced header with border
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        head_cell = ws.cell(
            row=start_row, column=stats_start_col, value="📊 Key Statistics (10Y)"
        )
        head_cell.font = Font(bold=True, color="FFFFFF", size=11)
        head_cell.fill = PatternFill("solid", fgColor=self.colors["header"])
        head_cell.alignment = Alignment(horizontal="center")
        head_cell.border = thin_border

        # Merge header cells
        ws.merge_cells(
            start_row=start_row,
            start_column=stats_start_col,
            end_row=start_row,
            end_column=stats_start_col + 1,
        )

        # Stats with enhanced formatting
        stats_data = [
            (
                "Maximum:",
                max_val,
                self.colors["positive"],
                "000000",
            ),  # Green bg, black text
            (
                "Average:",
                avg_val,
                self.colors["neutral"],
                "000000",
            ),  # Orange bg, black text
            (
                "Minimum:",
                min_val,
                self.colors["negative"],
                "FFFFFF",
            ),  # Red bg, white text
        ]

        # Add capacity only if available
        if capacity_val is not None:
            stats_data.append(
                ("Capacity:", capacity_val, self.colors["header"], "FFFFFF")
            )

        for i, (label, value, bg_color, text_color) in enumerate(stats_data, 1):
            label_cell = ws.cell(row=start_row + i, column=stats_start_col, value=label)
            label_cell.font = Font(
                bold=True, size=10, color="000000"
            )  # Black text for labels
            label_cell.border = thin_border

            value_cell = ws.cell(
                row=start_row + i, column=stats_start_col + 1, value=value
            )
            value_cell.number_format = "#,##0"
            value_cell.font = Font(bold=True, size=10, color=text_color)
            value_cell.fill = PatternFill("solid", fgColor=bg_color, fill_type="solid")
            value_cell.alignment = Alignment(horizontal="right")
            value_cell.border = thin_border

        ws.column_dimensions[stats_col_let].width = 14
        ws.column_dimensions[get_column_letter(stats_start_col + 1)].width = 12

        # --- ADD AVERAGE STORAGE TABLE BELOW KEY STATS ---
        table_start_row = start_row + len(stats_data) + 2

        # Calculate average storage by period for the table
        index_col_table = "Week_Num" if freq == "Weekly" else "Month"
        avg_storage_table = (
            df.groupby(index_col_table)["Storage_Level"].mean().reset_index()
        )
        avg_storage_table.columns = ["Period", "Avg_Storage"]

        # Table header
        table_header = ws.cell(
            row=table_start_row, column=stats_start_col, value=f"📊 Avg {freq} Storage"
        )
        table_header.font = Font(bold=True, color="FFFFFF", size=11)
        table_header.fill = PatternFill("solid", fgColor=self.colors["header"])
        table_header.alignment = Alignment(horizontal="center")
        table_header.border = thin_border
        ws.merge_cells(
            start_row=table_start_row,
            start_column=stats_start_col,
            end_row=table_start_row,
            end_column=stats_start_col + 1,
        )

        # Column headers
        ws.cell(
            row=table_start_row + 1, column=stats_start_col, value="Period"
        ).font = Font(bold=True, size=9)
        ws.cell(
            row=table_start_row + 1, column=stats_start_col, value="Period"
        ).fill = PatternFill("solid", fgColor=self.colors["subheader"])
        ws.cell(
            row=table_start_row + 1, column=stats_start_col, value="Period"
        ).font = Font(bold=True, color="FFFFFF", size=9)
        ws.cell(
            row=table_start_row + 1, column=stats_start_col, value="Period"
        ).border = thin_border

        ws.cell(
            row=table_start_row + 1, column=stats_start_col + 1, value="BCF"
        ).font = Font(bold=True, size=9)
        ws.cell(
            row=table_start_row + 1, column=stats_start_col + 1, value="BCF"
        ).fill = PatternFill("solid", fgColor=self.colors["subheader"])
        ws.cell(
            row=table_start_row + 1, column=stats_start_col + 1, value="BCF"
        ).font = Font(bold=True, color="FFFFFF", size=9)
        ws.cell(
            row=table_start_row + 1, column=stats_start_col + 1, value="BCF"
        ).border = thin_border
        ws.cell(
            row=table_start_row + 1, column=stats_start_col + 1, value="BCF"
        ).alignment = Alignment(horizontal="center")

        # Data rows
        for idx, row in avg_storage_table.iterrows():
            data_row = table_start_row + 2 + idx

            # Period label
            if freq == "Monthly":
                period_label = calendar.month_abbr[int(row["Period"])]
            else:
                period_label = int(row["Period"])

            period_cell = ws.cell(
                row=data_row, column=stats_start_col, value=period_label
            )
            period_cell.font = Font(bold=True, size=9, color="000000")
            period_cell.border = thin_border

            # Storage value
            storage_cell = ws.cell(
                row=data_row, column=stats_start_col + 1, value=row["Avg_Storage"]
            )
            storage_cell.number_format = "#,##0"
            storage_cell.font = Font(size=9, color="000000")
            storage_cell.alignment = Alignment(horizontal="right")
            storage_cell.border = thin_border

            # Alternating colors
            if idx % 2 == 0:
                period_cell.fill = PatternFill(
                    "solid", fgColor=self.colors["background"]
                )
                storage_cell.fill = PatternFill(
                    "solid", fgColor=self.colors["background"]
                )

        # --- 2. WRITE PIVOT TABLE ---
        ws.cell(
            row=start_row - 1,
            column=pivot_start_col,
            value="📈 Seasonality Analysis - Storage Levels",
        ).font = Font(bold=True, size=11, color="000000")

        years = pivot.columns.tolist()

        # Header row with enhanced styling
        period_cell = ws.cell(row=start_row, column=pivot_start_col, value="Period")
        period_cell.font = Font(
            bold=True, color="FFFFFF", size=10
        )  # White text on dark header
        period_cell.fill = PatternFill("solid", fgColor=self.colors["header"])
        period_cell.alignment = Alignment(horizontal="center")
        period_cell.border = thin_border

        for i, year in enumerate(years):
            cell = ws.cell(row=start_row, column=pivot_start_col + 1 + i, value=year)
            cell.font = Font(
                bold=True, color="FFFFFF", size=10
            )  # White text on dark header
            cell.fill = PatternFill("solid", fgColor=self.colors["subheader"])
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Data rows with alternating background
        for r_idx, (idx_val, row_data) in enumerate(pivot.iterrows()):
            row_num = start_row + 1 + r_idx

            # Period label
            period_cell = ws.cell(row=row_num, column=pivot_start_col, value=idx_val)
            period_cell.font = Font(bold=True, size=9, color="000000")  # Black text
            period_cell.border = thin_border

            if r_idx % 2 == 0:
                period_cell.fill = PatternFill(
                    "solid", fgColor=self.colors["background"]
                )

            # Data values
            for c_idx, val in enumerate(row_data):
                cell = ws.cell(row=row_num, column=pivot_start_col + 1 + c_idx)
                if pd.notna(val):
                    cell.value = val
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right")
                    cell.font = Font(color="000000")  # Black text
                cell.border = thin_border

                if r_idx % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=self.colors["background"])

        # Enhanced Color Grading
        first_data_row = start_row + 1
        last_data_row = start_row + len(pivot)
        first_col_let = get_column_letter(pivot_start_col + 1)
        last_col_let = get_column_letter(pivot_start_col + len(years))
        data_range = f"{first_col_let}{first_data_row}:{last_col_let}{last_data_row}"

        color_scale = ColorScaleRule(
            start_type="percentile",
            start_value=10,
            start_color="F8696B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="percentile",
            end_value=90,
            end_color="63BE7B",
        )
        ws.conditional_formatting.add(data_range, color_scale)

        # --- 3. CREATE MATPLOTLIB CHART ---
        fig, ax = plt.subplots(figsize=(14, 7), facecolor="white")
        ax.set_facecolor("#F8F9FA")

        # Plot each year with enhanced styling
        colors_palette = plt.cm.tab20(np.linspace(0, 1, len(years)))

        for idx, year in enumerate(years):
            if year in pivot.columns:
                data = pivot[year].dropna()
                if len(data) > 0:
                    # Emphasize the most recent year
                    if idx == len(years) - 1:
                        ax.plot(
                            data.index,
                            data.values,
                            linewidth=4,
                            label=str(year),
                            color="#C00000",
                            marker="o",
                            markersize=6,
                            markevery=2,
                            markeredgecolor="white",
                            markeredgewidth=1.5,
                            zorder=10,
                        )
                    else:
                        ax.plot(
                            data.index,
                            data.values,
                            linewidth=2,
                            label=str(year),
                            color=colors_palette[idx],
                            alpha=0.7,
                            marker=".",
                            markersize=4,
                            markevery=3,
                            zorder=5,
                        )

        # Formatting
        ax.set_title(
            f"Seasonality Pattern: {freq} Storage Levels (10-Year Comparison)",
            fontsize=16,
            fontweight="bold",
            pad=20,
            color="#1F4E78",
        )
        ax.set_xlabel(
            "Month" if freq == "Monthly" else "Week Number",
            fontsize=13,
            fontweight="bold",
            color="#333333",
        )
        ax.set_ylabel("Storage (BCF)", fontsize=13, fontweight="bold", color="#333333")

        # Format y-axis with commas
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f"{int(x):,}"))
        ax.tick_params(axis="both", which="major", labelsize=10, colors="#333333")

        # Enhanced grid
        ax.grid(
            True, alpha=0.4, linestyle="--", linewidth=0.8, color="#CCCCCC", zorder=1
        )
        ax.set_axisbelow(True)

        # Add border around plot area
        for spine in ax.spines.values():
            spine.set_edgecolor("#CCCCCC")
            spine.set_linewidth(1.5)

        # Legend - place outside plot area with better styling
        legend = ax.legend(
            loc="center left",
            bbox_to_anchor=(1, 0.5),
            frameon=True,
            shadow=True,
            fontsize=9,
            ncol=1,
            fancybox=True,
            framealpha=0.95,
            edgecolor="#CCCCCC",
        )
        legend.get_frame().set_facecolor("white")

        # Tight layout
        plt.tight_layout()

        # Save to bytes
        img_buffer = io.BytesIO()
        plt.savefig(
            img_buffer,
            format="png",
            dpi=150,
            bbox_inches="tight",
            facecolor="white",
            edgecolor="none",
        )
        img_buffer.seek(0)
        plt.close()

        # Insert into Excel - positioned below the pivot table
        chart_row = last_data_row + 3
        chart_col = pivot_start_col
        img = XLImage(img_buffer)
        img.width = 840
        img.height = 420
        ws.add_image(img, f"{get_column_letter(chart_col)}{chart_row}")

        return chart_row + 24  # Return the row after this chart for next charts

    def finalize_formatting_with_excel(self, output_file):
        """
        Applies specific Data Bar formatting:
        - Corrected Blue #5B9BD5 (Positive) / Orange #ED7D31 (Negative)
        - Axis forced to Dead Center
        - Vertical Axis Line HIDDEN (set to White)
        - No borders around bars
        """
        print(f"   ├─ ⚙️  Applying advanced color formatting (Blue/Orange, No Line)...")
        import win32com.client
        import os

        abs_path = os.path.abspath(output_file)

        try:
            # 1. Launch Excel
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False

            # 2. Open Workbook
            wb = excel.Workbooks.Open(abs_path)

            # 3. Define Colors (Corrected Calculation)
            # Formula: Red + (Green * 256) + (Blue * 65536)

            # Blue #5B9BD5 (R=91, G=155, B=213)
            # 91 + 39680 + 13959168 = 13998939
            color_blue = 13998939

            # Orange #ED7D31 (R=237, G=125, B=49)
            # 237 + 32000 + 3211264 = 3243501
            color_orange = 3243501

            # White #FFFFFF (for hiding the axis line)
            color_white = 16777215

            # 4. Apply Formatting
            for ws in wb.Worksheets:
                if "Weekly" in ws.Name or "Monthly" in ws.Name:
                    last_row = ws.Cells(ws.Rows.Count, "E").End(-4162).Row

                    if last_row > 2:
                        rng = ws.Range(f"E2:E{last_row}")

                        # Clear old rules
                        rng.FormatConditions.Delete()

                        # Add Data Bar
                        db = rng.FormatConditions.AddDatabar()

                        # --- COLORS ---
                        # Positive Bar (Blue Gradient)
                        db.BarColor.Color = color_blue
                        db.BarFillType = 1  # Gradient

                        # Negative Bar (Orange)
                        db.NegativeBarFormat.ColorType = 0  # Use specific color
                        db.NegativeBarFormat.Color.Color = color_orange
                        db.NegativeBarFormat.Color.TintAndShade = 0

                        # --- AXIS & BORDERS ---
                        # Force Axis to Dead Center
                        db.AxisPosition = 1  # 1 = xlDataBarAxisMidpoint

                        # Hide the Axis Line (Set color to White)
                        db.AxisColor.Color = color_white

                        # Remove Bar Borders
                        db.BarBorder.Type = 0  # 0 = xlDataBarBorderNone
                        db.NegativeBarFormat.BorderColorType = 0

            # 5. Finalize
            wb.Save()
            excel.Visible = True
            excel.DisplayAlerts = True
            print(f"   └─ ✓ Formatting applied. Excel file is now open.")

        except Exception as e:
            print(f"   ⚠️ Failed to automate Excel: {e}")
            try:
                excel.Quit()
            except:
                pass

    def format_worksheet(self, ws, df, freq, region_name="Total_Lower_48"):
        """Enhanced worksheet formatting with better visual hierarchy."""

        # Enhanced headers
        headers = [
            "📅 Date",
            "📆 Year",
            "📍 Period",
            "📦 Storage",
            "🔄 Net Flow",
            "📊 5Yr Avg Flow",
            "📈 Diff Flow",
            "📊 5Yr Avg Storage",
            "📈 Diff Storage",
            "🔁 YoY Flow",
        ]

        header_font = Font(
            bold=True, color="FFFFFF", size=11
        )  # White text for dark header
        header_fill = PatternFill("solid", fgColor=self.colors["header"])
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="medium"),
        )

        for idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = thin_border

        # Data rows with enhanced formatting
        data_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="hair"),
            bottom=Side(style="hair"),
        )

        for r_idx, row in df.iterrows():
            r = r_idx + 2

            # Alternating row colors
            row_fill = None
            if r_idx % 2 == 0:
                row_fill = PatternFill("solid", fgColor=self.colors["background"])

            # Date
            cell = ws.cell(r, 1, row["Week_Ending"])
            cell.number_format = "mm/dd/yyyy"
            cell.border = data_border
            cell.font = Font(color="000000")  # Black text
            if row_fill:
                cell.fill = row_fill

            # Year
            cell = ws.cell(r, 2, row["Year"])
            cell.border = data_border
            cell.font = Font(color="000000")  # Black text
            if row_fill:
                cell.fill = row_fill

            # Period
            if freq == "Monthly" and "Month" in row:
                period_val = calendar.month_name[int(row["Month"])]
            else:
                period_val = row["Week_Num"] if "Week_Num" in row else row["Month"]
            cell = ws.cell(r, 3, period_val)
            cell.border = data_border
            cell.font = Font(color="000000")  # Black text
            if row_fill:
                cell.fill = row_fill

            # Numeric values
            vals = [
                row.get(k)
                for k in [
                    "Storage_Level",
                    "Net_Flow",
                    "5Yr_Avg_Flow",
                    "Diff_vs_5Yr_Flow",
                    "5Yr_Avg_Storage",
                    "Diff_vs_5Yr_Storage",
                    "YoY_Change_Flow",
                ]
            ]

            for c, val in enumerate(vals, 4):
                cell = ws.cell(r, c)
                cell.border = data_border
                if row_fill:
                    cell.fill = row_fill

                if pd.notna(val):
                    cell.value = val
                    cell.number_format = "#,##0.0"
                    cell.alignment = Alignment(horizontal="right")

                    # Color code difference columns with contrast-based text color
                    if c in [7, 9, 10]:  # Diff columns
                        # Threshold: Values beyond +/- 60 get White text, others get Black
                        contrast_threshold = 60

                        if val >= contrast_threshold:
                            # Dark Green Background -> White Text
                            cell.font = Font(color="FFFFFF", bold=True)
                        elif val <= -contrast_threshold:
                            # Dark Red Background -> White Text
                            cell.font = Font(color="FFFFFF", bold=True)
                        else:
                            # Light Red/Green/Zero -> Black Text
                            cell.font = Font(color="000000", bold=True)
                    else:
                        cell.font = Font(color="000000")  # Black text for other columns

        # Enhanced conditional formatting with data bars
        rule = ColorScaleRule(
            start_type="num",
            start_value=-100,
            start_color="C00000",  # Red
            mid_type="num",
            mid_value=0,
            mid_color="FFFFFF",  # White
            end_type="num",
            end_value=100,
            end_color="70AD47",  # Green
        )
        ws.conditional_formatting.add(f"G2:G{ws.max_row}", rule)
        ws.conditional_formatting.add(f"I2:I{ws.max_row}", rule)
        ws.conditional_formatting.add(f"J2:J{ws.max_row}", rule)

        # --- Corrected Data Bar Rule ---
        # Locate the section where you add the flow_rule (around line 560)
        # Replace the existing flow_rule block with this:

        # We use 'min' and 'max' so Excel automatically calculates the Zero Axis.
        # This ensures Negative numbers grow Right-to-Left and Positives Left-to-Right.
        # flow_rule = DataBarRule(
        #     start_type='min',
        #     end_type='max',
        #     color="5B9BD5",  # Blue (Positive)
        #     showValue=True
        # )
        # ws.conditional_formatting.add(f"E2:E{ws.max_row}", flow_rule)

        # --- FIXED HEADER ROW ---
        ws.freeze_panes = "A2"

        # Auto-resize
        self.auto_resize_columns(ws)

        # Add seasonality chart and get the row position after it
        next_chart_row = self.add_seasonality_chart(ws, df, freq, region_name)

        # Add other charts below the seasonality chart in a single column
        # Column J is the 10th column (YoY Flow).
        # Column N (14) aligns nicely with the Seasonality Stats header.
        chart_col = 14
        next_chart_row = self.add_moving_average_chart(
            ws, df, next_chart_row, chart_col
        )
        next_chart_row = self.add_injection_withdrawal_chart(
            ws, df, freq, next_chart_row, chart_col
        )
        next_chart_row = self.add_average_storage_chart(
            ws, df, freq, next_chart_row, chart_col
        )

    def run_analysis(self, output_file):
        print("=" * 60)
        print("Starting ENHANCED Natural Gas Storage Analysis...")
        print("=" * 60)

        self.load_and_clean_data()
        wb = Workbook()
        wb.remove(wb.active)

        priority = [
            "Total_Lower_48",
            "Salt",
            "Non_Salt",
            "East",
            "Midwest",
            "South_Central",
            "Mountain",
            "Pacific",
        ]

        for region in priority:
            if region not in self.regions:
                continue

            raw_col = self.regions[region]
            print(f"\n📊 Processing {region}...")

            # Weekly analysis
            print(f"   ├─ Analyzing weekly data...")
            df_w = self.analyze_dataset(self.df_master, raw_col, "Weekly")
            ws_w = wb.create_sheet(f"Weekly_{region}")
            self.format_worksheet(ws_w, df_w, "Weekly", region)
            print(f"   ├─ ✓ Weekly sheet created")

            # Monthly analysis
            print(f"   ├─ Analyzing monthly data...")
            df_m_raw = self.create_monthly_data(self.df_master, raw_col)
            df_m = self.analyze_dataset(df_m_raw, "Storage_Level", "Monthly")
            ws_m = wb.create_sheet(f"Monthly_{region}")
            self.format_worksheet(ws_m, df_m, "Monthly", region)
            print(f"   └─ ✓ Monthly sheet created")

        print("\n" + "=" * 60)
        print(f"💾 Saving workbook to: {output_file}")
        wb.save(output_file)

        # This re-opens the saved file and fixes the colors automatically
        self.finalize_formatting_with_excel(output_file)

        print("=" * 60)
        print("✅ Analysis complete! File saved successfully.")
        print("=" * 60)


def main():
    input_file = r"ngshistory.xls"
    output_file = r"Natural_Gas_Master_Analysis_Enhanced.xlsx"

    analyzer = NaturalGasStorageAnalyzer(input_file)
    analyzer.run_analysis(output_file)


if __name__ == "__main__":
    main()
